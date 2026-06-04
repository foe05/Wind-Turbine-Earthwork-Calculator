"""Tests Wave 11: park_optimizer, site_data, multi_site_report."""

from pathlib import Path

import pytest
from shapely.geometry import Point

from app.core.earthwork import CutFillResult
from app.core.multi_site_report import export_multisite_xlsx, render_multisite_html
from app.core.multi_surface import MultiSurfaceResult, SurfaceType
from app.core.park_optimizer import (
    ParkOptimizer,
    SiteCandidate,
    SiteEarthwork,
    SiteWithCandidates,
    TransportConfig,
    euclidean_distance_km,
)
from app.core.site_data import MultiSiteProject, SiteData


# --------------------------------------------------------------- park optimizer LP

def test_park_lp_two_sites_simple_transport():
    """Site A hat Cut-Überschuss, Site B hat Fill-Bedarf — LP soll Material verschieben."""
    cfg = TransportConfig(
        cost_per_m3_km=0.20,
        dump_cost_per_m3=8.0,
        external_gravel_cost_per_m3=15.0,
    )
    sites = [
        SiteEarthwork("A", 0, 0, cut_excess_m3=100.0, fill_need_m3=0.0),
        SiteEarthwork("B", 1000, 0, cut_excess_m3=0.0, fill_need_m3=100.0),
    ]
    opt = ParkOptimizer(cfg)
    sol = opt.solve(sites)
    assert sol.solver_status.startswith(("ok", "Optim"))
    # Transport sollte ausgelöst werden: 1 km × 100 m³ × 0.20 = 20 €/m³ insgesamt = 20 € transport
    # Ersparnis = 100 * (8 + 15) - 20 = 2300 - 20 = 2280
    assert sol.savings_eur > 0
    assert len(sol.flows) == 1
    assert sol.flows[0].volume_m3 == pytest.approx(100.0, abs=0.1)


def test_park_lp_max_distance_blocks_transport():
    cfg = TransportConfig(
        cost_per_m3_km=0.20,
        dump_cost_per_m3=8.0,
        external_gravel_cost_per_m3=15.0,
        max_distance_km=0.5,  # nur < 500m erlaubt
    )
    sites = [
        SiteEarthwork("A", 0, 0, cut_excess_m3=100.0),
        SiteEarthwork("B", 1000, 0, fill_need_m3=100.0),  # 1 km entfernt
    ]
    sol = ParkOptimizer(cfg).solve(sites)
    # Kein Transport möglich → alles auf Baseline (Dump + Gravel)
    assert len(sol.flows) == 0
    assert sol.savings_eur == pytest.approx(0.0, abs=0.1)


def test_park_milp_chooses_low_cost_candidate():
    """MILP soll bei zwei Sites jeweils Kandidat mit niedrigeren Site-Kosten wählen."""
    cfg = TransportConfig(
        cost_per_m3_km=0.20,
        dump_cost_per_m3=8.0,
        external_gravel_cost_per_m3=15.0,
    )
    sites = [
        SiteWithCandidates("A", 0, 0, candidates=[
            SiteCandidate(cut_excess_m3=50, fill_need_m3=0, site_cost_eur=1000, label="hoch"),
            SiteCandidate(cut_excess_m3=20, fill_need_m3=0, site_cost_eur=500, label="niedrig"),
        ]),
        SiteWithCandidates("B", 200, 0, candidates=[
            SiteCandidate(cut_excess_m3=0, fill_need_m3=20, site_cost_eur=300),
        ]),
    ]
    sol = ParkOptimizer(cfg).solve_milp(sites)
    assert sol.solver_status.startswith(("ok", "Optim"))
    # Niedrigerer Site-Kosten-Kandidat sollte gewählt werden
    assert sol.chosen_candidate["A"].site_cost_eur == 500


def test_euclidean_distance_km():
    class _S:
        def __init__(self, x, y):
            self.x = x
            self.y = y
    a = _S(0, 0)
    b = _S(3000, 4000)
    assert euclidean_distance_km(a, b) == pytest.approx(5.0)


# --------------------------------------------------------------- site data + multi-site

def _mk_site(site_id: str, cost: float = 50000) -> SiteData:
    res = MultiSurfaceResult(
        crane_optimum_height=104.5,
        fok=104.0,
        foundation_depth=3.0,
        gravel_thickness=0.5,
        surface_results={
            SurfaceType.CRANE_PAD: CutFillResult(104.0, 4000, 1500, 2500, 100, 110, 105, 2500),
            SurfaceType.FOUNDATION: CutFillResult(101.0, 600, 0, 200, 101, 105, 103, 200),
        },
    )
    return SiteData(
        site_id=site_id,
        site_name=f"WEA {site_id}",
        location=Point(0, 0),
        result=res,
        costs={"cost_total": cost},
    )


def test_multi_site_add_and_rank():
    p = MultiSiteProject("Park Test")
    p.add_site(_mk_site("01", cost=50000))
    p.add_site(_mk_site("02", cost=70000))
    p.add_site(_mk_site("03", cost=30000))
    assert p.site_count == 3
    ranked = p.rank_by("total_cost", reverse=True)
    assert ranked[0].site_id == "02"
    assert ranked[-1].site_id == "03"
    assert p.best_site("total_cost").site_id == "03"
    assert p.worst_site("total_cost").site_id == "02"


def test_multi_site_duplicate_id_rejected():
    p = MultiSiteProject("Park")
    p.add_site(_mk_site("01"))
    with pytest.raises(ValueError, match="existiert bereits"):
        p.add_site(_mk_site("01"))


def test_multi_site_summary():
    p = MultiSiteProject("Park")
    p.add_site(_mk_site("01", cost=50000))
    p.add_site(_mk_site("02", cost=70000))
    s = p.summary()
    assert s["site_count"] == 2
    assert s["total_cost_eur"] == 120000
    assert s["max_cost_site"] == "02"
    assert s["min_cost_site"] == "01"


# --------------------------------------------------------------- multi-site report

def test_multi_site_html_render(tmp_path):
    p = MultiSiteProject("Park Test")
    p.add_site(_mk_site("01"))
    p.add_site(_mk_site("02"))
    out = tmp_path / "park.html"
    render_multisite_html(p, str(out))
    html = out.read_text()
    assert "Park Test" in html
    assert "WEA 01" in html
    assert "WEA 02" in html


def test_multi_site_html_with_park_solution(tmp_path):
    p = MultiSiteProject("Park Test")
    p.add_site(_mk_site("A"))
    p.add_site(_mk_site("B"))

    cfg = TransportConfig(0.20, 8.0, 15.0)
    sites_e = [
        SiteEarthwork("A", 0, 0, cut_excess_m3=100),
        SiteEarthwork("B", 500, 0, fill_need_m3=100),
    ]
    sol = ParkOptimizer(cfg).solve(sites_e)

    out = tmp_path / "park.html"
    render_multisite_html(p, str(out), park_solution=sol)
    html = out.read_text()
    assert "Park-Optimierung (LP)" in html
    assert "Materialflüsse" in html


def test_multi_site_xlsx_export(tmp_path):
    p = MultiSiteProject("Park Excel")
    p.add_site(_mk_site("01", cost=50000))
    p.add_site(_mk_site("02", cost=70000))
    out = tmp_path / "park.xlsx"
    export_multisite_xlsx(p, str(out))
    assert out.exists()
    assert out.stat().st_size > 4000

    # Verifikation: kann openpyxl es wieder lesen?
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "Park-Übersicht" in wb.sheetnames
    assert "Standorte" in wb.sheetnames
    sheet = wb["Standorte"]
    assert sheet.cell(row=2, column=1).value == "01"
