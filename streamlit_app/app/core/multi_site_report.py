"""
Multi-Site Vergleichs-Report (HTML + XLSX).

Pragmatischer MVP-Port von core/multi_site_report_generator.py: tabellarische
Übersicht aller Sites, Park-Optimierung (optional), Best/Worst Highlights.
"""

from __future__ import annotations

import html
from datetime import datetime
from pathlib import Path
from typing import Optional

from .park_optimizer import ParkMILPSolution, ParkSolution
from .site_data import MultiSiteProject


def render_multisite_html(
    project: MultiSiteProject,
    output_path: str | Path,
    park_solution: Optional[ParkSolution | ParkMILPSolution] = None,
) -> str:
    """Renders Multi-Site-Report als HTML."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    title = html.escape(project.project_name)
    when = datetime.now().strftime("%Y-%m-%d %H:%M")
    summary = project.summary()

    rows = []
    for s in project.sites:
        rows.append(
            f"<tr>"
            f"<td>{html.escape(s.site_id)}</td>"
            f"<td>{html.escape(s.site_name)}</td>"
            f"<td class='r'>{s.crane_height:.2f}</td>"
            f"<td class='r'>{s.total_cut:.0f}</td>"
            f"<td class='r'>{s.total_fill:.0f}</td>"
            f"<td class='r'>{s.net_volume:.0f}</td>"
            f"<td class='r'>{s.total_cost:,.0f}</td>"
            f"</tr>"
        )
    rows_html = "\n".join(rows)

    park_section = ""
    if park_solution is not None:
        if isinstance(park_solution, ParkMILPSolution):
            chosen = "<br>".join(
                f"{html.escape(sid)}: Kandidat #{idx}"
                for sid, idx in park_solution.chosen_index.items()
            )
            park_section = f"""
<h2>Park-Optimierung (MILP)</h2>
<p>Status: {html.escape(park_solution.solver_status)}</p>
<p><b>Site-Kosten:</b> {park_solution.total_site_cost_eur:,.0f} €</p>
<p><b>Transport:</b> {park_solution.total_transport_eur:,.0f} €</p>
<p><b>Deponie:</b> {park_solution.total_dump_eur:,.0f} €</p>
<p><b>Schotter-Import:</b> {park_solution.total_gravel_eur:,.0f} €</p>
<p><b>Gesamt:</b> {park_solution.total_cost_eur:,.0f} €</p>
<h3>Gewählte Kandidaten</h3><p>{chosen}</p>
<h3>Materialflüsse</h3>
<table><thead><tr><th>Von</th><th>Nach</th><th>Volumen [m³]</th><th>Distanz [km]</th><th>Kosten [€]</th></tr></thead><tbody>
{"".join(f"<tr><td>{html.escape(f.from_site)}</td><td>{html.escape(f.to_site)}</td><td class='r'>{f.volume_m3:.0f}</td><td class='r'>{f.distance_km:.2f}</td><td class='r'>{f.transport_cost_eur:,.0f}</td></tr>" for f in park_solution.flows)}
</tbody></table>"""
        else:
            park_section = f"""
<h2>Park-Optimierung (LP)</h2>
<p>Status: {html.escape(park_solution.solver_status)}</p>
<p><b>Transport:</b> {park_solution.total_transport_eur:,.0f} €</p>
<p><b>Restliche Deponie:</b> {park_solution.total_dump_eur:,.0f} €</p>
<p><b>Restl. Schotter-Import:</b> {park_solution.total_gravel_eur:,.0f} €</p>
<p><b>Baseline (ohne Optimierung):</b> {park_solution.baseline_cost_eur:,.0f} €</p>
<p><b>Ersparnis:</b> {park_solution.savings_eur:,.0f} €</p>
<h3>Materialflüsse</h3>
<table><thead><tr><th>Von</th><th>Nach</th><th>Volumen [m³]</th><th>Distanz [km]</th><th>Kosten [€]</th></tr></thead><tbody>
{"".join(f"<tr><td>{html.escape(f.from_site)}</td><td>{html.escape(f.to_site)}</td><td class='r'>{f.volume_m3:.0f}</td><td class='r'>{f.distance_km:.2f}</td><td class='r'>{f.transport_cost_eur:,.0f}</td></tr>" for f in park_solution.flows)}
</tbody></table>"""

    h = f"""<!DOCTYPE html>
<html lang="de"><head><meta charset="UTF-8">
<title>{title} — Multi-Site Vergleich</title>
<style>
  body {{ font-family: Arial, sans-serif; color:#222; margin:24px; }}
  h1 {{ font-size:22px; }} h2 {{ font-size:17px; border-bottom:1px solid #ccc; padding-bottom:4px; }}
  table {{ border-collapse:collapse; width:100%; margin:8px 0 16px 0; font-size:13px; }}
  th, td {{ border:1px solid #ddd; padding:5px 8px; }}
  th {{ background:#f5f5f5; text-align:left; }}
  td.r {{ text-align:right; }}
  .totals td {{ background:#eef6fc; font-weight:bold; }}
</style></head><body>
<h1>{title} — Multi-Site Vergleich</h1>
<p>Erstellt am {when} · {summary['site_count']} Standorte</p>

<h2>Park-Übersicht</h2>
<table>
  <tr><th>Standorte</th><td class='r'>{summary['site_count']}</td></tr>
  <tr><th>Gesamt-Abtrag</th><td class='r'>{summary['total_cut_m3']:,.0f} m³</td></tr>
  <tr><th>Gesamt-Auftrag</th><td class='r'>{summary['total_fill_m3']:,.0f} m³</td></tr>
  <tr><th>Netto-Bilanz</th><td class='r'>{summary['total_net_volume_m3']:,.0f} m³</td></tr>
  <tr><th>Gesamt-Kosten</th><td class='r'>{summary['total_cost_eur']:,.0f} €</td></tr>
  <tr><th>Ø Abtrag/Site</th><td class='r'>{summary['avg_cut_per_site_m3']:,.0f} m³</td></tr>
  <tr><th>Ø Kosten/Site</th><td class='r'>{summary['avg_cost_per_site_eur']:,.0f} €</td></tr>
  <tr><th>Günstigster Standort</th><td class='r'>{html.escape(str(summary.get('min_cost_site') or ''))}</td></tr>
  <tr><th>Teuerster Standort</th><td class='r'>{html.escape(str(summary.get('max_cost_site') or ''))}</td></tr>
</table>

<h2>Standorte je WEA</h2>
<table><thead><tr>
<th>ID</th><th>Name</th><th>Kran [m ü.NN]</th><th>Cut [m³]</th><th>Fill [m³]</th><th>Netto [m³]</th><th>Kosten [€]</th>
</tr></thead><tbody>
{rows_html}
</tbody></table>

{park_section}
</body></html>"""
    output_path.write_text(h, encoding="utf-8")
    return str(output_path)


def export_multisite_xlsx(project: MultiSiteProject, output_path: str | Path) -> str:
    """Schreibt eine XLSX-Datei mit Park-Summary + Site-Tabelle."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Park-Übersicht"
    bold = Font(bold=True)

    summary = project.summary()
    summary_sheet["A1"] = project.project_name
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet["A3"] = "Standorte"
    summary_sheet["B3"] = summary["site_count"]
    summary_sheet["A4"] = "Gesamt-Abtrag [m³]"
    summary_sheet["B4"] = summary["total_cut_m3"]
    summary_sheet["A5"] = "Gesamt-Auftrag [m³]"
    summary_sheet["B5"] = summary["total_fill_m3"]
    summary_sheet["A6"] = "Netto-Bilanz [m³]"
    summary_sheet["B6"] = summary["total_net_volume_m3"]
    summary_sheet["A7"] = "Gesamt-Kosten [€]"
    summary_sheet["B7"] = summary["total_cost_eur"]
    for cell in ("A3", "A4", "A5", "A6", "A7"):
        summary_sheet[cell].font = bold

    sites_sheet = wb.create_sheet("Standorte")
    headers = ["Site-ID", "Name", "Kran [m ü.NN]", "Cut [m³]", "Fill [m³]", "Netto [m³]", "Kosten [€]"]
    for col, h in enumerate(headers, start=1):
        c = sites_sheet.cell(row=1, column=col, value=h)
        c.font = bold
    for row_idx, s in enumerate(project.sites, start=2):
        sites_sheet.cell(row=row_idx, column=1, value=s.site_id)
        sites_sheet.cell(row=row_idx, column=2, value=s.site_name)
        sites_sheet.cell(row=row_idx, column=3, value=round(s.crane_height, 2))
        sites_sheet.cell(row=row_idx, column=4, value=round(s.total_cut, 1))
        sites_sheet.cell(row=row_idx, column=5, value=round(s.total_fill, 1))
        sites_sheet.cell(row=row_idx, column=6, value=round(s.net_volume, 1))
        sites_sheet.cell(row=row_idx, column=7, value=round(s.total_cost, 2))

    wb.save(output_path)
    return str(output_path)
