
import unittest
from unittest.mock import MagicMock, patch, mock_open
import os
import sys
import tempfile
from pathlib import Path

# Add parent directory to path to import modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Mock weasyprint and openpyxl before importing MultiSiteReportGenerator
sys.modules['weasyprint'] = MagicMock()
sys.modules['openpyxl'] = MagicMock()
sys.modules['openpyxl.styles'] = MagicMock()
sys.modules['openpyxl.utils'] = MagicMock()

from windturbine_earthwork_calculator_v2.core.multi_site_report_generator import MultiSiteReportGenerator


class TestMultiSiteReportGenerator(unittest.TestCase):
    def setUp(self):
        """Set up test fixtures."""
        # Create mock site results
        self.site_1 = {
            'site_id': 'WEA01',
            'site_name': 'Turbine 01',
            'coordinates': (500000.0, 5800000.0),
            'results': {
                'total_cut': 1000.0,
                'total_fill': 800.0,
                'net_volume': 200.0,
                'gravel_fill_external': 50.0,
                'crane_height': 100.0,
                'platform_height': 100.0,
                'terrain_min': 95.0,
                'terrain_max': 105.0,
                'terrain_mean': 100.0,
                'platform_area': 1500.0,
                'total_platform_area': 1500.0,
                'total_area': 2000.0,
                'slope_width': 5.0,
                'platform_cut': 500.0,
                'platform_fill': 400.0,
                'slope_cut': 500.0,
                'slope_fill': 400.0,
                'surfaces': {}
            },
            'config': {
                'slope_angle': 45.0
            }
        }

        self.site_2 = {
            'site_id': 'WEA02',
            'site_name': 'Turbine 02',
            'coordinates': (510000.0, 5810000.0),
            'results': {
                'total_cut': 1500.0,
                'total_fill': 1200.0,
                'net_volume': 300.0,
                'gravel_fill_external': 75.0,
                'crane_height': 105.0,
                'platform_height': 105.0,
                'terrain_min': 98.0,
                'terrain_max': 112.0,
                'terrain_mean': 105.0,
                'platform_area': 1600.0,
                'total_platform_area': 1600.0,
                'total_area': 2200.0,
                'slope_width': 5.5,
                'platform_cut': 750.0,
                'platform_fill': 600.0,
                'slope_cut': 750.0,
                'slope_fill': 600.0,
                'surfaces': {}
            },
            'config': {
                'slope_angle': 45.0
            }
        }

        self.site_3 = {
            'site_id': 'WEA03',
            'site_name': 'Turbine 03',
            'coordinates': (520000.0, 5820000.0),
            'results': {
                'total_cut': 800.0,
                'total_fill': 600.0,
                'net_volume': 200.0,
                'gravel_fill_external': 40.0,
                'crane_height': 98.0,
                'platform_height': 98.0,
                'terrain_min': 92.0,
                'terrain_max': 104.0,
                'terrain_mean': 98.0,
                'platform_area': 1450.0,
                'total_platform_area': 1450.0,
                'total_area': 1900.0,
                'slope_width': 4.8,
                'platform_cut': 400.0,
                'platform_fill': 300.0,
                'slope_cut': 400.0,
                'slope_fill': 300.0,
                'surfaces': {}
            },
            'config': {
                'slope_angle': 45.0
            }
        }

        self.site_results = [self.site_1, self.site_2, self.site_3]

        self.cost_config = {
            'cut_cost_per_m3': 5.0,
            'fill_cost_per_m3': 8.0,
            'gravel_cost_per_m3': 25.0,
            'transport_cost_per_m3_km': 0.5
        }

    def test_initialization_with_empty_sites(self):
        """Test initialization with empty site list."""
        generator = MultiSiteReportGenerator([], self.cost_config)

        self.assertEqual(generator.total_cut, 0)
        self.assertEqual(generator.total_fill, 0)
        self.assertEqual(generator.total_volume_moved, 0)
        self.assertEqual(generator.total_net_volume, 0)
        self.assertEqual(generator.total_gravel, 0)
        self.assertEqual(generator.total_cost, 0)
        self.assertEqual(generator.avg_cut, 0)
        self.assertEqual(generator.avg_fill, 0)
        self.assertEqual(generator.min_cut, 0)
        self.assertEqual(generator.max_cut, 0)
        self.assertEqual(generator.min_fill, 0)
        self.assertEqual(generator.max_fill, 0)

    def test_initialization_with_single_site(self):
        """Test initialization with single site."""
        generator = MultiSiteReportGenerator([self.site_1], self.cost_config)

        self.assertEqual(generator.total_cut, 1000.0)
        self.assertEqual(generator.total_fill, 800.0)
        self.assertEqual(generator.total_volume_moved, 1800.0)
        self.assertEqual(generator.total_net_volume, 200.0)
        self.assertEqual(generator.total_gravel, 50.0)
        self.assertEqual(generator.avg_cut, 1000.0)
        self.assertEqual(generator.avg_fill, 800.0)
        self.assertEqual(generator.min_cut, 1000.0)
        self.assertEqual(generator.max_cut, 1000.0)
        self.assertEqual(generator.min_fill, 800.0)
        self.assertEqual(generator.max_fill, 800.0)

    def test_initialization_with_multiple_sites(self):
        """Test initialization with multiple sites."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        # Total cut: 1000 + 1500 + 800 = 3300
        self.assertEqual(generator.total_cut, 3300.0)
        # Total fill: 800 + 1200 + 600 = 2600
        self.assertEqual(generator.total_fill, 2600.0)
        # Total volume moved: 1800 + 2700 + 1400 = 5900
        self.assertEqual(generator.total_volume_moved, 5900.0)
        # Total net volume: 200 + 300 + 200 = 700
        self.assertEqual(generator.total_net_volume, 700.0)
        # Total gravel: 50 + 75 + 40 = 165
        self.assertEqual(generator.total_gravel, 165.0)
        # Averages
        self.assertAlmostEqual(generator.avg_cut, 1100.0, places=2)
        self.assertAlmostEqual(generator.avg_fill, 866.67, places=2)
        # Min/Max
        self.assertEqual(generator.min_cut, 800.0)
        self.assertEqual(generator.max_cut, 1500.0)
        self.assertEqual(generator.min_fill, 600.0)
        self.assertEqual(generator.max_fill, 1200.0)

    def test_calculate_site_cost(self):
        """Test site cost calculation."""
        generator = MultiSiteReportGenerator([self.site_1], self.cost_config)

        results = self.site_1['results']
        cost = generator._calculate_site_cost(results)

        # Expected cost:
        # Cut: 1000 * 5 = 5000
        # Fill: 800 * 8 = 6400
        # Gravel: 50 * 25 = 1250
        # Transport: (1000 + 800) * 0.5 * 5 = 4500
        # Total: 5000 + 6400 + 1250 + 4500 = 17150
        self.assertAlmostEqual(cost, 17150.0, places=2)

    def test_default_cost_config(self):
        """Test default cost configuration is applied."""
        generator = MultiSiteReportGenerator([self.site_1])

        # Check default values
        self.assertEqual(generator.cost_config['cut_cost_per_m3'], 5.0)
        self.assertEqual(generator.cost_config['fill_cost_per_m3'], 8.0)
        self.assertEqual(generator.cost_config['gravel_cost_per_m3'], 25.0)
        self.assertEqual(generator.cost_config['transport_cost_per_m3_km'], 0.5)

    def test_calculated_cost_added_to_sites(self):
        """Test that calculated cost is added to each site."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        for site in generator.site_results:
            self.assertIn('calculated_cost', site)
            self.assertGreater(site['calculated_cost'], 0)

        # Verify total cost is sum of individual costs
        total = sum(site['calculated_cost'] for site in generator.site_results)
        self.assertAlmostEqual(generator.total_cost, total, places=2)

    def test_generate_header(self):
        """Test HTML header generation."""
        generator = MultiSiteReportGenerator([self.site_1], self.cost_config)

        header = generator._generate_header("Test Project")

        self.assertIn('Multi-Site Erdmassenvergleich', header)
        self.assertIn('Test Project', header)
        self.assertIn('Erstellt am:', header)
        self.assertIn('class="header"', header)

    def test_generate_summary(self):
        """Test summary section generation."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        summary = generator._generate_summary()

        self.assertIn('Projektzusammenfassung', summary)
        self.assertIn('3', summary)  # Number of sites
        self.assertIn('Windenergieanlagen-Standorte', summary)
        self.assertIn('Gesamt Abtrag', summary)
        self.assertIn('Gesamt Auftrag', summary)
        self.assertIn('Gesamt Erdbewegungen', summary)
        self.assertIn('Netto-Bilanz', summary)
        self.assertIn('Externes Schottermaterial', summary)
        # Check values are present (as formatted strings)
        self.assertIn('3,300', summary)  # Total cut
        self.assertIn('2,600', summary)  # Total fill

    def test_generate_statistics(self):
        """Test statistics section generation."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        statistics = generator._generate_statistics()

        self.assertIn('Statistische Auswertung', statistics)
        self.assertIn('Abtrag-Statistik', statistics)
        self.assertIn('Auftrag-Statistik', statistics)
        self.assertIn('Durchschnitt', statistics)
        self.assertIn('Minimum', statistics)
        self.assertIn('Maximum', statistics)

    def test_generate_site_ranking(self):
        """Test site ranking section generation."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        ranking = generator._generate_site_ranking()

        self.assertIn('Standort-Rangliste nach Komplexität', ranking)
        self.assertIn('Turbine 02', ranking)  # Highest volume (2700)
        self.assertIn('Turbine 01', ranking)  # Second (1800)
        self.assertIn('Turbine 03', ranking)  # Lowest (1400)
        self.assertIn('rank-badge', ranking)
        self.assertIn('Empfehlung', ranking)

    def test_site_ranking_order(self):
        """Test sites are ranked in descending order by volume moved."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        ranking = generator._generate_site_ranking()

        # Check that Turbine 02 (highest) appears before Turbine 03 (lowest)
        pos_site2 = ranking.find('Turbine 02')
        pos_site3 = ranking.find('Turbine 03')
        self.assertLess(pos_site2, pos_site3)

    def test_generate_site_comparison(self):
        """Test site comparison section generation."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        comparison = generator._generate_site_comparison()

        self.assertIn('Detaillierter Standortvergleich', comparison)
        self.assertIn('Turbine 01', comparison)
        self.assertIn('Turbine 02', comparison)
        self.assertIn('Turbine 03', comparison)
        self.assertIn('Koordinaten', comparison)
        self.assertIn('Kranstellflächen-Höhe', comparison)
        self.assertIn('Höhenunterschied', comparison)
        self.assertIn('Plattformfläche', comparison)

    def test_generate_site_details(self):
        """Test individual site details section generation."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        details = generator._generate_site_details()

        self.assertIn('Detaillierte Standort-Einzelauswertung', details)
        self.assertIn('Turbine 01', details)
        self.assertIn('Volumenübersicht', details)
        self.assertIn('Geländestatistik', details)
        self.assertIn('Volumenaufschlüsselung nach Komponente', details)
        self.assertIn('Kostenaufschlüsselung', details)
        self.assertIn('Plattformkonfiguration', details)
        self.assertIn('Standortkoordinaten', details)

    def test_generate_cost_breakdown(self):
        """Test cost breakdown section generation."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        breakdown = generator._generate_cost_breakdown()

        self.assertIn('Kostenaufschlüsselung', breakdown)
        self.assertIn('Kostenkalkulations-Parameter', breakdown)
        self.assertIn('Abtrag-Kosten', breakdown)
        self.assertIn('Auftrag-Kosten', breakdown)
        self.assertIn('Schotter-Kosten', breakdown)
        self.assertIn('Transport-Kosten', breakdown)
        self.assertIn('Gesamtkosten', breakdown)
        self.assertIn('5.00 €/m³', breakdown)  # Cut cost parameter
        self.assertIn('8.00 €/m³', breakdown)  # Fill cost parameter

    def test_generate_footer(self):
        """Test footer generation."""
        generator = MultiSiteReportGenerator([self.site_1], self.cost_config)

        footer = generator._generate_footer()

        self.assertIn('Multi-Site Erdmassenvergleich', footer)
        self.assertIn('QGIS Processing Plugin', footer)
        self.assertIn('Bericht erstellt am:', footer)
        self.assertIn('class="footer"', footer)

    def test_get_css_styles(self):
        """Test CSS styles generation."""
        generator = MultiSiteReportGenerator([self.site_1], self.cost_config)

        css = generator._get_css_styles()

        self.assertIn('<style>', css)
        self.assertIn('</style>', css)
        self.assertIn('.header', css)
        self.assertIn('.section', css)
        self.assertIn('.card', css)
        self.assertIn('.rank-badge', css)
        self.assertIn('@media print', css)

    def test_generate_html_creates_file(self):
        """Test HTML generation creates a valid file."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'test_report.html')
            generator.generate_html(output_path, project_name='Test Wind Farm')

            # Verify file was created
            self.assertTrue(os.path.exists(output_path))

            # Verify file content
            with open(output_path, 'r', encoding='utf-8') as f:
                content = f.read()

            self.assertIn('<!DOCTYPE html>', content)
            self.assertIn('Test Wind Farm', content)
            self.assertIn('Projektzusammenfassung', content)
            self.assertIn('Standort-Rangliste', content)
            self.assertIn('Turbine 01', content)
            self.assertIn('Turbine 02', content)
            self.assertIn('Turbine 03', content)

    def test_generate_html_with_default_project_name(self):
        """Test HTML generation with default project name."""
        generator = MultiSiteReportGenerator([self.site_1], self.cost_config)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'test_report.html')
            generator.generate_html(output_path)

            with open(output_path, 'r', encoding='utf-8') as f:
                content = f.read()

            self.assertIn('Windpark-Projekt', content)

    def test_generate_pdf(self):
        """Test PDF generation calls WeasyPrint correctly."""
        generator = MultiSiteReportGenerator([self.site_1], self.cost_config)

        with tempfile.TemporaryDirectory() as tmpdir:
            html_path = os.path.join(tmpdir, 'test.html')
            pdf_path = os.path.join(tmpdir, 'test.pdf')

            # Create a dummy HTML file
            Path(html_path).write_text('<html><body>Test</body></html>', encoding='utf-8')

            # Mock weasyprint to verify it's called correctly
            with patch('windturbine_earthwork_calculator_v2.core.multi_site_report_generator.HTML') as mock_html:
                mock_html_instance = MagicMock()
                mock_html.return_value = mock_html_instance

                generator.generate_pdf(html_path, pdf_path)

                # Verify HTML was called with correct path
                mock_html.assert_called_once()
                call_args = mock_html.call_args
                self.assertIn('filename', call_args[1])

                # Verify write_pdf was called
                mock_html_instance.write_pdf.assert_called_once()

    @unittest.skip("Requires openpyxl to be installed")
    def test_generate_excel_creates_file(self):
        """Test Excel generation creates a valid file."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'test_report.xlsx')
            generator.generate_excel(output_path, project_name='Test Wind Farm')

            # Verify file was created
            self.assertTrue(os.path.exists(output_path))

    @unittest.skip("Requires openpyxl to be installed")
    def test_generate_excel_has_required_sheets(self):
        """Test Excel file has all required sheets."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'test_report.xlsx')
            generator.generate_excel(output_path)

            # Read the Excel file and verify sheets
            from openpyxl import load_workbook
            wb = load_workbook(output_path)

            self.assertIn('Summary', wb.sheetnames)
            self.assertIn('Sites Ranking', wb.sheetnames)
            self.assertIn('Individual Sites', wb.sheetnames)

    @unittest.skip("Requires openpyxl to be installed")
    def test_create_summary_sheet(self):
        """Test summary sheet creation in Excel."""
        from openpyxl import Workbook

        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)
        wb = Workbook()

        generator._create_summary_sheet(wb, 'Test Project')

        ws = wb['Summary']
        # Check title
        self.assertEqual(ws['A1'].value, 'Multi-Site Erdmassenvergleich')
        self.assertEqual(ws['A2'].value, 'Test Project')

        # Verify sheet has content (spot check a few cells)
        found_site_count = False
        found_total_cost = False
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value == 'Anzahl Standorte:':
                    found_site_count = True
                if cell.value == 'Gesamtkosten (geschätzt):':
                    found_total_cost = True

        self.assertTrue(found_site_count, "Site count label not found in summary sheet")
        self.assertTrue(found_total_cost, "Total cost label not found in summary sheet")

    @unittest.skip("Requires openpyxl to be installed")
    def test_create_ranking_sheet(self):
        """Test ranking sheet creation in Excel."""
        from openpyxl import Workbook

        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)
        wb = Workbook()

        generator._create_ranking_sheet(wb)

        ws = wb['Sites Ranking']
        # Check title
        self.assertEqual(ws['A1'].value, 'Standort-Rangliste nach Komplexität')

        # Verify headers exist
        found_rang = False
        found_standort = False
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value == 'Rang':
                    found_rang = True
                if cell.value == 'Standort':
                    found_standort = True

        self.assertTrue(found_rang, "Rang header not found")
        self.assertTrue(found_standort, "Standort header not found")

    @unittest.skip("Requires openpyxl to be installed")
    def test_create_individual_sites_sheet(self):
        """Test individual sites sheet creation in Excel."""
        from openpyxl import Workbook

        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)
        wb = Workbook()

        generator._create_individual_sites_sheet(wb)

        ws = wb['Individual Sites']
        # Check title
        self.assertEqual(ws['A1'].value, 'Detaillierte Standort-Einzelauswertung')

        # Verify site names are present
        found_turbine_01 = False
        found_turbine_02 = False
        for row in ws.iter_rows(min_row=1, max_row=100):
            for cell in row:
                if cell.value and 'Turbine 01' in str(cell.value):
                    found_turbine_01 = True
                if cell.value and 'Turbine 02' in str(cell.value):
                    found_turbine_02 = True

        self.assertTrue(found_turbine_01, "Turbine 01 not found in individual sites sheet")
        self.assertTrue(found_turbine_02, "Turbine 02 not found in individual sites sheet")

    def test_site_details_with_new_surface_structure(self):
        """Test site details generation with new multi-surface structure."""
        # Create site with new surface structure
        site_with_surfaces = {
            'site_id': 'WEA04',
            'site_name': 'Turbine 04',
            'coordinates': (530000.0, 5830000.0),
            'results': {
                'total_cut': 2000.0,
                'total_fill': 1500.0,
                'net_volume': 500.0,
                'gravel_fill_external': 100.0,
                'crane_height': 110.0,
                'terrain_min': 100.0,
                'terrain_max': 115.0,
                'terrain_mean': 107.5,
                'platform_area': 1800.0,
                'total_platform_area': 1800.0,
                'total_area': 2500.0,
                'slope_width': 6.0,
                'surfaces': {
                    'kranstellflaeche': {
                        'cut': 900.0,
                        'fill': 700.0
                    },
                    'fundamentflaeche': {
                        'cut': 300.0,
                        'fill': 200.0
                    },
                    'auslegerflaeche': {
                        'cut': 500.0,
                        'fill': 400.0
                    },
                    'rotorflaeche': {
                        'cut': 300.0,
                        'fill': 200.0
                    }
                }
            },
            'config': {
                'slope_angle': 45.0
            }
        }

        generator = MultiSiteReportGenerator([site_with_surfaces], self.cost_config)
        details = generator._generate_site_details()

        self.assertIn('Turbine 04', details)
        self.assertIn('Kranstellfläche', details)
        self.assertIn('Weitere Flächen', details)

    def test_generate_html_all_sections_present(self):
        """Test that generated HTML contains all expected sections."""
        generator = MultiSiteReportGenerator(self.site_results, self.cost_config)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'test_report.html')
            generator.generate_html(output_path, project_name='Complete Test')

            with open(output_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # Check all major sections are present
            sections = [
                'Projektzusammenfassung',
                'Statistische Auswertung',
                'Standort-Rangliste nach Komplexität',
                'Detaillierter Standortvergleich',
                'Detaillierte Standort-Einzelauswertung',
                'Kostenaufschlüsselung'
            ]

            for section in sections:
                self.assertIn(section, content, f"Section '{section}' not found in HTML")

    def test_empty_site_results_handles_gracefully(self):
        """Test that empty results dict is handled correctly."""
        site_with_empty_results = {
            'site_id': 'EMPTY',
            'site_name': 'Empty Site',
            'coordinates': (0, 0),
            'results': {},
            'config': {}
        }

        generator = MultiSiteReportGenerator([site_with_empty_results], self.cost_config)

        # Should not crash
        self.assertEqual(generator.total_cut, 0)
        self.assertEqual(generator.total_fill, 0)

    def test_missing_optional_fields(self):
        """Test that missing optional fields are handled correctly."""
        minimal_site = {
            'site_id': 'MIN',
            'site_name': 'Minimal Site',
            'results': {
                'total_cut': 100.0,
                'total_fill': 80.0
            }
        }

        generator = MultiSiteReportGenerator([minimal_site], self.cost_config)

        # Should use defaults for missing fields
        summary = generator._generate_summary()
        self.assertIn('Projektzusammenfassung', summary)


if __name__ == '__main__':
    unittest.main()
